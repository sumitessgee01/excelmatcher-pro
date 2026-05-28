from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .normalizer import normalize_barcode

try:
    import xlsxwriter  # noqa: F401

    HAS_XLSXWRITER = True
except Exception:
    HAS_XLSXWRITER = False


STATUS_PRIORITY_SHEET12 = {
    "Brand Mismatch": 0,
    "Value Mismatch": 1,
    "Qty Mismatch": 2,
    "Not In Brand": 3,
    "Not In EssGee": 3,
    "Not In Data": 3,
    "Match": 9,
    # Legacy statuses
    "Mismatch": 1,
    "Only In F1": 3,
    "Only In F2": 3,
    "Matched": 9,
}

STATUS_PRIORITY_SHEET3 = {
    "Qty Mismatch": 0,
    "Brand Mismatch": 1,
    "Value Mismatch": 2,
    "Not In Brand": 3,
    "Not In EssGee": 3,
    "Not In Data": 3,
    "Match": 9,
    # Legacy statuses
    "Mismatch": 2,
    "Only In F1": 3,
    "Only In F2": 3,
    "Matched": 9,
}

STATUS_STYLES = {
    "Match": {"fill": "C6EFCE", "font": "276221"},
    "Brand Mismatch": {"fill": "FFC7CE", "font": "9C0006"},
    "Value Mismatch": {"fill": "FFC7CE", "font": "9C0006"},
    "Qty Mismatch": {"fill": "FFEB9C", "font": "9C6500"},
    "Not In Brand": {"fill": "BDD7EE", "font": "1F497D"},
    "Not In EssGee": {"fill": "BDD7EE", "font": "1F497D"},
    "Not In Data": {"fill": "BDD7EE", "font": "1F497D"},
    # Legacy statuses
    "Matched": {"fill": "C6EFCE", "font": "276221"},
    "Mismatch": {"fill": "FFC7CE", "font": "9C0006"},
    "Only In F1": {"fill": "BDD7EE", "font": "1F497D"},
    "Only In F2": {"fill": "BDD7EE", "font": "1F497D"},
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FILL = PatternFill(fill_type="solid", fgColor="0F172A")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
MISMATCH_CELL_FILL = PatternFill(fill_type="solid", fgColor="FF0000")
MISMATCH_CELL_FONT = Font(color="FFFFFF", bold=True)
QTY_MISMATCH_FILL = PatternFill(fill_type="solid", fgColor="FF6600")
QTY_MISMATCH_FONT = Font(color="FFFFFF", bold=True)
ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
FAST_STYLE_ROW_THRESHOLD = 12000
MISMATCH_HIGHLIGHT_ROW_LIMIT = 20000
NUMBER_FORMAT_ROW_LIMIT = 40000
ULTRA_FAST_ROW_THRESHOLD = 20000

STATUS_FILL_CACHE = {
    key: PatternFill(fill_type="solid", fgColor=value["fill"])
    for key, value in STATUS_STYLES.items()
}
STATUS_FONT_CACHE = {
    key: Font(color=value["font"], bold=False)
    for key, value in STATUS_STYLES.items()
}


@dataclass(slots=True)
class RowMeta:
    status: str
    match_remark: str
    detailed_remark: str
    mismatch_labels: set[str]


class ExcelExporter:
    def export_summary(
        self,
        output_path: str | Path,
        brand_df: pd.DataFrame,
        essgee_df: pd.DataFrame,
        rows: list[dict[str, Any]],
        *,
        key_mappings: list[dict[str, Any]] | None = None,
        value_mappings: list[dict[str, Any]] | None = None,
        brand_file_name: str,
        essgee_file_name: str,
        progress_callback: Callable[[int, str], None] | None = None,
    ) -> Path:
        output = Path(output_path).expanduser().resolve()
        output.parent.mkdir(parents=True, exist_ok=True)

        self._progress(progress_callback, 20, "Preparing summary export data")
        brand_meta = self._build_side_meta(rows, side="f1")
        essgee_meta = self._build_side_meta(rows, side="f2")
        brand_out = self._build_configured_output_df(
            brand_df,
            brand_meta,
            side="f1",
            key_mappings=key_mappings or [],
            value_mappings=value_mappings or [],
        )
        essgee_out = self._build_configured_output_df(
            essgee_df,
            essgee_meta,
            side="f2",
            key_mappings=key_mappings or [],
            value_mappings=value_mappings or [],
        )
        ultra_fast = max(len(brand_out), len(essgee_out), len(rows)) >= ULTRA_FAST_ROW_THRESHOLD

        mismatch_out = brand_out[~brand_out["Match Status"].isin(["Matched", "Match"])].copy()
        if not ultra_fast:
            mismatch_out = self._sort_by_status(mismatch_out, STATUS_PRIORITY_SHEET3)

        brand_write = brand_out.drop(columns=["_mismatch_labels"], errors="ignore")
        essgee_write = essgee_out.drop(columns=["_mismatch_labels"], errors="ignore")
        mismatch_write = mismatch_out.drop(columns=["_mismatch_labels"], errors="ignore")
        engine, engine_kwargs, use_xlsxwriter_fast = self._writer_settings(ultra_fast)
        mode_text = "xlsxwriter-fast" if use_xlsxwriter_fast else "openpyxl-standard"
        self._progress(progress_callback, 32, f"Export engine: {mode_text}")
        writer_kwargs: dict[str, Any] = {"engine": engine}
        if engine_kwargs:
            writer_kwargs["engine_kwargs"] = engine_kwargs

        self._progress(progress_callback, 45, "Writing summary sheets")
        with pd.ExcelWriter(output, **writer_kwargs) as writer:
            if use_xlsxwriter_fast:
                ws_brand = writer.book.add_worksheet("Brand Data")
                ws_essgee = writer.book.add_worksheet("Ess Gee Data")
                ws_mismatch = writer.book.add_worksheet("Brand Mismatch Data")
                writer.sheets["Brand Data"] = ws_brand
                writer.sheets["Ess Gee Data"] = ws_essgee
                writer.sheets["Brand Mismatch Data"] = ws_mismatch
                self._write_df_stream_xlsxwriter(ws_brand, brand_write, start_row=0)
                self._write_df_stream_xlsxwriter(ws_essgee, essgee_write, start_row=0)
                self._write_df_stream_xlsxwriter(ws_mismatch, mismatch_write, start_row=7)
            else:
                brand_write.to_excel(writer, sheet_name="Brand Data", index=False)
                essgee_write.to_excel(writer, sheet_name="Ess Gee Data", index=False)
                mismatch_write.to_excel(writer, sheet_name="Brand Mismatch Data", index=False, startrow=7)
            self._progress(progress_callback, 65, "Applying sheet formatting")
            if use_xlsxwriter_fast:
                fmts = self._build_xlsxwriter_formats(writer.book)
                self._style_clean_sheet_xlsxwriter(
                    writer.sheets["Brand Data"],
                    brand_write,
                    fmts,
                )
                self._style_clean_sheet_xlsxwriter(
                    writer.sheets["Ess Gee Data"],
                    essgee_write,
                    fmts,
                )
                self._style_mismatch_sheet_xlsxwriter(
                    writer.sheets["Brand Mismatch Data"],
                    mismatch_write,
                    fmts,
                    brand_file_name=brand_file_name,
                    essgee_file_name=essgee_file_name,
                    start_row=7,
                )
            else:
                ws_brand = writer.sheets["Brand Data"]
                ws_essgee = writer.sheets["Ess Gee Data"]
                ws_mismatch = writer.sheets["Brand Mismatch Data"]
                self._style_clean_sheet(
                    ws_brand,
                    df=brand_out,
                    meta=brand_meta,
                    value_mappings=value_mappings or [],
                    qty_col=None,
                    side="f1",
                    title="Brand Data",
                    fast_mode=ultra_fast,
                )
                self._style_clean_sheet(
                    ws_essgee,
                    df=essgee_out,
                    meta=essgee_meta,
                    value_mappings=value_mappings or [],
                    qty_col=None,
                    side="f2",
                    title="Ess Gee Data",
                    fast_mode=ultra_fast,
                )
                self._style_mismatch_sheet(
                    ws_mismatch,
                    mismatch_write,
                    brand_file_name=brand_file_name,
                    essgee_file_name=essgee_file_name,
                    fast_mode=ultra_fast,
                )
        self._progress(progress_callback, 95, "Summary export finalized")
        return output

    def export_full(
        self,
        output_path: str | Path,
        brand_df: pd.DataFrame,
        essgee_df: pd.DataFrame,
        rows: list[dict[str, Any]],
        *,
        brand_file_name: str,
        essgee_file_name: str,
        value_mappings: list[dict[str, Any]] | None = None,
        qty_f1_col: str | None = None,
        qty_f2_col: str | None = None,
        progress_callback: Callable[[int, str], None] | None = None,
    ) -> Path:
        output = Path(output_path).expanduser().resolve()
        output.parent.mkdir(parents=True, exist_ok=True)

        self._progress(progress_callback, 20, "Preparing full export data")
        brand_meta = self._build_side_meta(rows, side="f1")
        essgee_meta = self._build_side_meta(rows, side="f2")

        brand_out = self._build_output_df(brand_df, brand_meta, side="f1")
        essgee_out = self._build_output_df(essgee_df, essgee_meta, side="f2")
        ultra_fast = max(len(brand_out), len(essgee_out), len(rows)) >= ULTRA_FAST_ROW_THRESHOLD
        if not ultra_fast:
            brand_out = self._sort_by_status(brand_out, STATUS_PRIORITY_SHEET12)
            essgee_out = self._sort_by_status(essgee_out, STATUS_PRIORITY_SHEET12)
        brand_write = brand_out.drop(columns=["_mismatch_labels"], errors="ignore")
        essgee_write = essgee_out.drop(columns=["_mismatch_labels"], errors="ignore")

        mismatch_out = brand_out[~brand_out["Match Status"].isin(["Matched", "Match"])].copy()
        if not ultra_fast:
            mismatch_out = self._sort_by_status(mismatch_out, STATUS_PRIORITY_SHEET3)
        mismatch_out = mismatch_out[
            ["Match Status", "Match Remark", "Detailed Remark"]
            + [
                c
                for c in brand_out.columns
                if c not in {"Match Status", "Match Remark", "Detailed Remark", "_mismatch_labels"}
            ]
        ]
        engine, engine_kwargs, use_xlsxwriter_fast = self._writer_settings(ultra_fast)
        mode_text = "xlsxwriter-fast" if use_xlsxwriter_fast else "openpyxl-standard"
        self._progress(progress_callback, 32, f"Export engine: {mode_text}")
        writer_kwargs: dict[str, Any] = {"engine": engine}
        if engine_kwargs:
            writer_kwargs["engine_kwargs"] = engine_kwargs

        self._progress(progress_callback, 45, "Writing full export sheets")
        with pd.ExcelWriter(output, **writer_kwargs) as writer:
            if use_xlsxwriter_fast:
                ws_brand = writer.book.add_worksheet("Brand File (Cleaned)")
                ws_essgee = writer.book.add_worksheet("EssGee File (Cleaned)")
                ws_mismatch = writer.book.add_worksheet("Brand Mismatch Report")
                writer.sheets["Brand File (Cleaned)"] = ws_brand
                writer.sheets["EssGee File (Cleaned)"] = ws_essgee
                writer.sheets["Brand Mismatch Report"] = ws_mismatch
                self._write_df_stream_xlsxwriter(ws_brand, brand_write, start_row=0)
                self._write_df_stream_xlsxwriter(ws_essgee, essgee_write, start_row=0)
                self._write_df_stream_xlsxwriter(ws_mismatch, mismatch_out, start_row=7)
            else:
                brand_write.to_excel(writer, sheet_name="Brand File (Cleaned)", index=False)
                essgee_write.to_excel(writer, sheet_name="EssGee File (Cleaned)", index=False)
                mismatch_out.to_excel(writer, sheet_name="Brand Mismatch Report", index=False, startrow=7)
            self._progress(progress_callback, 65, "Applying full export formatting")
            if use_xlsxwriter_fast:
                fmts = self._build_xlsxwriter_formats(writer.book)
                self._style_clean_sheet_xlsxwriter(
                    writer.sheets["Brand File (Cleaned)"],
                    brand_write,
                    fmts,
                )
                self._style_clean_sheet_xlsxwriter(
                    writer.sheets["EssGee File (Cleaned)"],
                    essgee_write,
                    fmts,
                )
                self._style_mismatch_sheet_xlsxwriter(
                    writer.sheets["Brand Mismatch Report"],
                    mismatch_out,
                    fmts,
                    brand_file_name=brand_file_name,
                    essgee_file_name=essgee_file_name,
                    start_row=7,
                )
            else:
                ws_brand = writer.sheets["Brand File (Cleaned)"]
                ws_essgee = writer.sheets["EssGee File (Cleaned)"]
                ws_mismatch = writer.sheets["Brand Mismatch Report"]

                self._style_clean_sheet(
                    ws_brand,
                    df=brand_out,
                    meta=brand_meta,
                    value_mappings=value_mappings or [],
                    qty_col=qty_f1_col,
                    side="f1",
                    fast_mode=ultra_fast,
                )
                self._style_clean_sheet(
                    ws_essgee,
                    df=essgee_out,
                    meta=essgee_meta,
                    value_mappings=value_mappings or [],
                    qty_col=qty_f2_col,
                    side="f2",
                    fast_mode=ultra_fast,
                )

                self._style_mismatch_sheet(
                    ws_mismatch,
                    mismatch_out,
                    brand_file_name=brand_file_name,
                    essgee_file_name=essgee_file_name,
                    fast_mode=ultra_fast,
                )
        self._progress(progress_callback, 95, "Full export finalized")
        return output

    def export_mismatch_only(
        self,
        output_path: str | Path,
        brand_df: pd.DataFrame,
        rows: list[dict[str, Any]],
        *,
        brand_file_name: str,
        essgee_file_name: str,
        progress_callback: Callable[[int, str], None] | None = None,
    ) -> Path:
        output = Path(output_path).expanduser().resolve()
        output.parent.mkdir(parents=True, exist_ok=True)

        self._progress(progress_callback, 20, "Preparing mismatch export data")
        brand_meta = self._build_side_meta(rows, side="f1")
        brand_out = self._build_output_df(brand_df, brand_meta, side="f1")
        ultra_fast = max(len(brand_out), len(rows)) >= ULTRA_FAST_ROW_THRESHOLD
        mismatch_out = brand_out[~brand_out["Match Status"].isin(["Matched", "Match"])].copy()
        if not ultra_fast:
            mismatch_out = self._sort_by_status(mismatch_out, STATUS_PRIORITY_SHEET3)
        mismatch_out = mismatch_out[
            ["Match Status", "Match Remark", "Detailed Remark"]
            + [
                c
                for c in brand_out.columns
                if c not in {"Match Status", "Match Remark", "Detailed Remark", "_mismatch_labels"}
            ]
        ]
        engine, engine_kwargs, use_xlsxwriter_fast = self._writer_settings(ultra_fast)
        mode_text = "xlsxwriter-fast" if use_xlsxwriter_fast else "openpyxl-standard"
        self._progress(progress_callback, 32, f"Export engine: {mode_text}")
        writer_kwargs: dict[str, Any] = {"engine": engine}
        if engine_kwargs:
            writer_kwargs["engine_kwargs"] = engine_kwargs

        self._progress(progress_callback, 45, "Writing mismatch export sheet")
        with pd.ExcelWriter(output, **writer_kwargs) as writer:
            if use_xlsxwriter_fast:
                ws = writer.book.add_worksheet("Brand Mismatch Report")
                writer.sheets["Brand Mismatch Report"] = ws
                self._write_df_stream_xlsxwriter(ws, mismatch_out, start_row=7)
            else:
                mismatch_out.to_excel(writer, sheet_name="Brand Mismatch Report", index=False, startrow=7)
            self._progress(progress_callback, 65, "Applying mismatch formatting")
            if use_xlsxwriter_fast:
                fmts = self._build_xlsxwriter_formats(writer.book)
                self._style_mismatch_sheet_xlsxwriter(
                    writer.sheets["Brand Mismatch Report"],
                    mismatch_out,
                    fmts,
                    brand_file_name=brand_file_name,
                    essgee_file_name=essgee_file_name,
                    start_row=7,
                )
            else:
                ws = writer.sheets["Brand Mismatch Report"]
                self._style_mismatch_sheet(
                    ws,
                    mismatch_out,
                    brand_file_name=brand_file_name,
                    essgee_file_name=essgee_file_name,
                    fast_mode=ultra_fast,
                )
        self._progress(progress_callback, 95, "Mismatch export finalized")
        return output

    def _progress(
        self,
        progress_callback: Callable[[int, str], None] | None,
        value: int,
        message: str,
    ) -> None:
        if progress_callback is None:
            return
        try:
            progress_callback(int(value), message)
        except Exception:
            return

    def _writer_settings(self, ultra_fast: bool) -> tuple[str, dict[str, Any] | None, bool]:
        if ultra_fast and HAS_XLSXWRITER:
            return "xlsxwriter", {"options": {"constant_memory": True, "strings_to_urls": False}}, True
        return "openpyxl", None, False

    def _to_xlsx_cell_value(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, str):
            return value
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
        # pandas/numpy scalars -> Python scalars
        if hasattr(value, "item"):
            try:
                value = value.item()
            except Exception:
                pass
        if hasattr(value, "to_pydatetime"):
            try:
                return value.to_pydatetime()
            except Exception:
                return str(value)
        return value

    def _write_df_stream_xlsxwriter(self, ws, df: pd.DataFrame, *, start_row: int = 0) -> None:
        if df is None:
            return
        headers = [str(c) for c in df.columns.tolist()]
        ws.write_row(start_row, 0, headers)
        row_no = start_row + 1
        for row in df.itertuples(index=False, name=None):
            ws.write_row(row_no, 0, [self._to_xlsx_cell_value(v) for v in row])
            row_no += 1

    def _build_xlsxwriter_formats(self, workbook) -> dict[str, Any]:
        fmts: dict[str, Any] = {
            "header": workbook.add_format(
                {
                    "bold": True,
                    "font_color": "#FFFFFF",
                    "bg_color": "#1F2937",
                    "align": "center",
                    "valign": "vcenter",
                }
            ),
            "title": workbook.add_format(
                {
                    "bold": True,
                    "font_color": "#FFFFFF",
                    "bg_color": "#0F172A",
                    "align": "center",
                    "valign": "vcenter",
                    "font_size": 14,
                }
            ),
            "meta": workbook.add_format({"bold": False, "font_color": "#0F172A"}),
        }
        for status, style in STATUS_STYLES.items():
            fmts[f"status::{status}"] = workbook.add_format(
                {
                    "font_color": f"#{style['font']}",
                    "bg_color": f"#{style['fill']}",
                }
            )
        return fmts

    def _style_clean_sheet_xlsxwriter(
        self,
        ws,
        df: pd.DataFrame,
        fmts: dict[str, Any],
    ) -> None:
        if df is None:
            return
        row_count = len(df)
        col_count = len(df.columns)
        if col_count <= 0:
            return

        ws.set_row(0, 22, fmts["header"])
        ws.freeze_panes(1, 0)
        if row_count > 0:
            ws.autofilter(0, 0, row_count, col_count - 1)

        for idx, width in enumerate(self._xlsxwriter_widths(df, sample_rows=40)):
            ws.set_column(idx, idx, width)

        self._apply_status_cf_xlsxwriter(
            ws,
            df,
            status_col_name="Match Status",
            first_data_row=1,
            last_data_row=row_count,
            color_only_status=True,
            fmts=fmts,
        )

    def _style_summary_sheet_xlsxwriter(
        self,
        ws,
        df: pd.DataFrame,
        fmts: dict[str, Any],
    ) -> None:
        if df is None:
            return
        row_count = len(df)
        col_count = len(df.columns)
        if col_count <= 0:
            return

        ws.set_row(0, 22, fmts["header"])
        ws.freeze_panes(1, 0)
        if row_count > 0:
            ws.autofilter(0, 0, row_count, col_count - 1)

        for idx, width in enumerate(self._xlsxwriter_widths(df, sample_rows=40)):
            ws.set_column(idx, idx, width)

        self._apply_status_cf_xlsxwriter(
            ws,
            df,
            status_col_name="Match Status",
            first_data_row=1,
            last_data_row=row_count,
            color_only_status=True,
            fmts=fmts,
        )

    def _style_mismatch_sheet_xlsxwriter(
        self,
        ws,
        mismatch_df: pd.DataFrame,
        fmts: dict[str, Any],
        *,
        brand_file_name: str,
        essgee_file_name: str,
        start_row: int = 7,
    ) -> None:
        # Title + metadata area.
        now = datetime.now().strftime("%d-%b-%Y %H:%M")
        total = len(mismatch_df)
        mismatch = int(
            mismatch_df["Match Status"].isin(["Mismatch", "Brand Mismatch", "Value Mismatch"]).sum()
        ) if "Match Status" in mismatch_df.columns else 0
        qty_mismatch = int((mismatch_df["Match Status"] == "Qty Mismatch").sum()) if "Match Status" in mismatch_df.columns else 0
        not_in_data = int(
            mismatch_df["Match Status"].isin(
                ["Only In F1", "Only In F2", "Not In Data", "Not In Brand", "Not In EssGee"]
            ).sum()
        ) if "Match Status" in mismatch_df.columns else 0

        col_count = len(mismatch_df.columns)
        merge_to = max(0, max(5, col_count - 1))
        ws.merge_range(0, 0, 0, merge_to, "BRAND MISMATCH REPORT", fmts["title"])
        ws.write(1, 0, f"Generated: {now}", fmts["meta"])
        ws.write(2, 0, f"Brand File: {brand_file_name}", fmts["meta"])
        ws.write(3, 0, f"EssGee File: {essgee_file_name}", fmts["meta"])
        ws.write(
            5,
            0,
            f"SUMMARY | Mismatch: {mismatch} | Qty Mismatch: {qty_mismatch} | Not In Data: {not_in_data} | Total: {total}",
            fmts["meta"],
        )

        # Header row written by pandas is at start_row.
        ws.set_row(start_row, 22, fmts["header"])
        ws.freeze_panes(start_row + 1, 0)
        if col_count > 0:
            ws.autofilter(start_row, 0, start_row + total, col_count - 1)
            for idx, width in enumerate(self._xlsxwriter_widths(mismatch_df, sample_rows=40)):
                ws.set_column(idx, idx, width)

        self._apply_status_cf_xlsxwriter(
            ws,
            mismatch_df,
            status_col_name="Match Status",
            first_data_row=start_row + 1,
            last_data_row=start_row + total,
            color_only_status=True,
            fmts=fmts,
        )

    def _apply_status_cf_xlsxwriter(
        self,
        ws,
        df: pd.DataFrame,
        *,
        status_col_name: str,
        first_data_row: int,
        last_data_row: int,
        color_only_status: bool,
        fmts: dict[str, Any],
    ) -> None:
        if first_data_row > last_data_row or df is None or len(df.columns) == 0:
            return
        if status_col_name not in df.columns:
            return
        status_col = int(df.columns.get_loc(status_col_name))
        first_col = status_col if color_only_status else 0
        last_col = status_col if color_only_status else (len(df.columns) - 1)

        # Use one conditional rule per status (no Python row loops).
        for status in (
            "Brand Mismatch",
            "Value Mismatch",
            "Qty Mismatch",
            "Not In Brand",
            "Not In EssGee",
            "Not In Data",
            "Match",
            "Mismatch",
            "Matched",
            "Only In F1",
            "Only In F2",
        ):
            fmt = fmts.get(f"status::{status}")
            if fmt is None:
                continue
            ws.conditional_format(
                first_data_row,
                first_col,
                last_data_row,
                last_col,
                {
                    "type": "cell",
                    "criteria": "==",
                    "value": f'"{status}"',
                    "format": fmt,
                },
            )

    def _xlsxwriter_widths(self, df: pd.DataFrame, sample_rows: int = 40) -> list[int]:
        if df is None or df.empty:
            return [12 for _ in df.columns] if df is not None else []
        sample = df.head(sample_rows)
        widths: list[int] = []
        for col in df.columns:
            base = len(str(col))
            try:
                observed = int(sample[col].astype(str).map(len).max() or 0)
            except Exception:
                observed = base
            w = max(base, observed)
            widths.append(max(10, min(60, int(w * 1.1) + 2)))
        return widths

    def _build_summary_df(
        self,
        rows: list[dict[str, Any]],
        *,
        key_mappings: list[dict[str, Any]],
        value_mappings: list[dict[str, Any]],
        sort_status: bool = True,
    ) -> pd.DataFrame:
        ordered_keys = self._ordered_key_mappings(key_mappings)
        ordered_values = self._ordered_value_mappings(value_mappings)
        ordered_mappings = ordered_keys + ordered_values

        mapping_labels = self._unique_labels([self._mapping_label(m) for m in ordered_mappings])
        base_cols = ["Match Status", "Match Remark", "Detailed Remark", "Mismatch Columns"]
        mapping_cols: list[str] = []
        for label in mapping_labels:
            mapping_cols.append(f"Brand {label}")
            mapping_cols.append(f"EssGee {label}")
        all_cols = base_cols + mapping_cols + ["_mismatch_labels"]

        records: list[dict[str, Any]] = []
        for row in rows:
            status = str(row.get("match_status", "") or "")
            summary_status = self._summary_status(status)

            f1_row = row.get("f1_row") or {}
            f2_row = row.get("f2_row") or {}
            mismatch_labels = [str(x) for x in (row.get("mismatch_columns") or []) if str(x).strip()]

            rec: dict[str, Any] = {
                "Match Status": summary_status,
                "Match Remark": str(row.get("match_remark", "") or ""),
                "Detailed Remark": str(row.get("detailed_remark", "") or ""),
                "Mismatch Columns": ", ".join(mismatch_labels),
                "_mismatch_labels": set(mismatch_labels),
            }

            for mapping, label in zip(ordered_mappings, mapping_labels):
                f1_col = str(mapping.get("f1_col", "") or "")
                f2_col = str(mapping.get("f2_col", "") or "")
                rec[f"Brand {label}"] = f1_row.get(f1_col) if f1_col else None
                rec[f"EssGee {label}"] = f2_row.get(f2_col) if f2_col else None

            records.append(rec)

        if not records:
            return pd.DataFrame(columns=all_cols)

        out = pd.DataFrame(records)
        for col in all_cols:
            if col not in out.columns:
                out[col] = None
        out = out[all_cols]
        out = self._coerce_barcode_columns_to_text(out)
        if sort_status:
            out = self._sort_by_status(out, STATUS_PRIORITY_SHEET3)
        return out

    def _build_configured_output_df(
        self,
        source_df: pd.DataFrame,
        meta: dict[int, RowMeta],
        *,
        side: str,
        key_mappings: list[dict[str, Any]],
        value_mappings: list[dict[str, Any]],
    ) -> pd.DataFrame:
        source = source_df.reset_index(drop=True).copy(deep=False)
        ordered_mappings = self._ordered_key_mappings(key_mappings) + self._ordered_value_mappings(value_mappings)
        side_col = "f1_col" if side == "f1" else "f2_col"

        out = pd.DataFrame(index=source.index)
        if ordered_mappings:
            labels = self._unique_labels([self._mapping_label(mapping) for mapping in ordered_mappings])
            for mapping, label in zip(ordered_mappings, labels):
                source_col = str(mapping.get(side_col, "") or "")
                if source_col and source_col in source.columns:
                    out[label] = source[source_col].values
                else:
                    out[label] = None
        else:
            # Defensive fallback for old saved jobs without mappings.
            for col in source.columns:
                out[str(col)] = source[col].values

        row_count = len(source)
        default_status, default_remark, default_detail = self._default_side_remarks(side)
        statuses = [default_status] * row_count
        remarks = [default_remark] * row_count
        details = [default_detail] * row_count
        mismatch_labels: list[set[str]] = [set() for _ in range(row_count)]

        for idx, row_meta in meta.items():
            if idx < 0 or idx >= row_count:
                continue
            statuses[idx] = row_meta.status
            remarks[idx] = row_meta.match_remark
            details[idx] = row_meta.detailed_remark
            mismatch_labels[idx] = row_meta.mismatch_labels

        out["Match Status"] = statuses
        out["Match Remark"] = remarks
        out["Detailed Remark"] = details
        out["_mismatch_labels"] = mismatch_labels
        out = self._coerce_barcode_columns_to_text(out)
        return out

    def _default_side_remarks(self, side: str) -> tuple[str, str, str]:
        if side == "f1":
            return "Not In EssGee", "Not In EssGee", "Invoice not found in EssGee data."
        return "Not In Brand", "Not In Brand", "Invoice not found in Brand data."

    def _ordered_key_mappings(self, mappings: list[dict[str, Any]]) -> list[dict[str, Any]]:
        def key_priority(item: dict[str, Any], idx: int) -> tuple[int, int]:
            text = self._mapping_text(item)
            if any(x in text for x in ("invoice", "bill", "inv")):
                return (0, idx)
            if any(x in text for x in ("party", "customer", "dealer", "vendor")):
                return (1, idx)
            if "brand" in text:
                return (2, idx)
            return (10, idx)

        indexed = [(idx, m) for idx, m in enumerate(mappings) if m]
        indexed.sort(key=lambda pair: key_priority(pair[1], pair[0]))
        return [m for _, m in indexed]

    def _ordered_value_mappings(self, mappings: list[dict[str, Any]]) -> list[dict[str, Any]]:
        def value_priority(item: dict[str, Any], idx: int) -> tuple[int, int]:
            text = self._mapping_text(item)
            if "mrp" in text and all(x not in text for x in ("value", "val", "amount")):
                return (0, idx)
            if any(x in text for x in ("qty", "quantity")):
                return (1, idx)
            if "mrp" in text and any(x in text for x in ("value", "val", "amount")):
                return (2, idx)
            if any(x in text for x in ("discount", "disc", "dis ")):
                return (3, idx)
            if "net" in text:
                return (4, idx)
            return (20, idx)

        indexed = [(idx, m) for idx, m in enumerate(mappings) if m]
        indexed.sort(key=lambda pair: value_priority(pair[1], pair[0]))
        return [m for _, m in indexed]

    def _mapping_text(self, mapping: dict[str, Any]) -> str:
        label = str(mapping.get("label", "") or "")
        f1 = str(mapping.get("f1_col", "") or "")
        f2 = str(mapping.get("f2_col", "") or "")
        return f"{label} {f1} {f2}".casefold()

    def _mapping_label(self, mapping: dict[str, Any]) -> str:
        label = str(mapping.get("label", "") or "").strip()
        if label:
            return label
        col = str(mapping.get("f1_col", mapping.get("f2_col", "Column")) or "Column").strip()
        return col or "Column"

    def _summary_status(self, status: str) -> str:
        if status == "Only In F1":
            return "Not In EssGee"
        if status == "Only In F2":
            return "Not In Brand"
        if status == "Matched":
            return "Match"
        if status == "Mismatch":
            return "Value Mismatch"
        return status or "Unknown"

    def _unique_labels(self, labels: list[str]) -> list[str]:
        counts: dict[str, int] = {}
        out: list[str] = []
        for label in labels:
            base = (label or "Column").strip() or "Column"
            n = counts.get(base, 0) + 1
            counts[base] = n
            out.append(base if n == 1 else f"{base} ({n})")
        return out

    def _build_side_meta(self, rows: list[dict[str, Any]], side: str) -> dict[int, RowMeta]:
        meta: dict[int, RowMeta] = {}
        index_field = "f1_index" if side == "f1" else "f2_index"
        for row in rows:
            idx = row.get(index_field)
            if idx is None:
                continue
            idx_int = int(idx)
            status = str(row.get("match_status", "") or "")
            remark = str(row.get("match_remark", "") or "")
            detail = str(row.get("detailed_remark", "") or "")
            labels = set((row.get("mismatch_columns") or []))

            existing = meta.get(idx_int)
            if not existing:
                meta[idx_int] = RowMeta(status=status, match_remark=remark, detailed_remark=detail, mismatch_labels=labels)
                continue

            old_p = STATUS_PRIORITY_SHEET12.get(existing.status, 99)
            new_p = STATUS_PRIORITY_SHEET12.get(status, 99)
            if new_p < old_p:
                meta[idx_int] = RowMeta(status=status, match_remark=remark, detailed_remark=detail, mismatch_labels=labels)
            else:
                existing.mismatch_labels.update(labels)
        return meta

    def _build_output_df(self, source_df: pd.DataFrame, meta: dict[int, RowMeta], side: str) -> pd.DataFrame:
        out = source_df.reset_index(drop=True).copy(deep=False)
        row_count = len(out)

        if side == "f1":
            default_status = "Not In EssGee"
            default_remark = "Not In EssGee"
            default_detail = "Invoice not found in EssGee data."
        else:
            default_status = "Not In Brand"
            default_remark = "Not In Brand"
            default_detail = "Invoice not found in Brand data."

        statuses = [default_status] * row_count
        remarks = [default_remark] * row_count
        details = [default_detail] * row_count
        mismatch_labels: list[set[str]] = [set() for _ in range(row_count)]

        for idx, row_meta in meta.items():
            if idx < 0 or idx >= row_count:
                continue
            statuses[idx] = row_meta.status
            remarks[idx] = row_meta.match_remark
            details[idx] = row_meta.detailed_remark
            mismatch_labels[idx] = row_meta.mismatch_labels

        out["Match Status"] = statuses
        out["Match Remark"] = remarks
        out["Detailed Remark"] = details
        out["_mismatch_labels"] = mismatch_labels
        out = self._coerce_barcode_columns_to_text(out)
        return out

    def _coerce_barcode_columns_to_text(self, df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            return df
        out = df.copy(deep=False)
        for col in out.columns:
            col_text = str(col or "").casefold()
            if not any(token in col_text for token in ("barcode", "ean", "upc")):
                continue
            out[col] = out[col].map(self._barcode_text_cell)
        return out

    def _barcode_text_cell(self, value: Any) -> Any:
        if value is None:
            return None
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
        normalized = normalize_barcode(value)
        if normalized:
            return normalized
        text = str(value).strip()
        return text if text else None

    def _sort_by_status(self, df: pd.DataFrame, priority_map: dict[str, int]) -> pd.DataFrame:
        out = df.copy()
        out["_status_priority"] = out["Match Status"].map(lambda x: priority_map.get(str(x), 99))
        out = out.sort_values(by=["_status_priority"], kind="stable").reset_index(drop=True)
        out.drop(columns=["_status_priority"], inplace=True)
        return out

    def _style_clean_sheet(
        self,
        ws,
        df: pd.DataFrame,
        meta: dict[int, RowMeta],
        value_mappings: list[dict[str, Any]],
        qty_col: str | None,
        side: str,
        title: str | None = None,
        fast_mode: bool = False,
    ) -> None:
        header_row = 1
        data_start_row = 2
        if title:
            ws.insert_rows(1)
            ws.cell(row=1, column=1).value = title
            ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="0F172A")
            header_row = 2
            data_start_row = 3

        headers = [cell.value for cell in ws[header_row]]
        for cell in ws[header_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER

        status_col = headers.index("Match Status") + 1 if "Match Status" in headers else None
        qty_col_idx = headers.index(qty_col) + 1 if qty_col and qty_col in headers else None
        remark_col = headers.index("Match Remark") + 1 if "Match Remark" in headers else None
        detail_col = headers.index("Detailed Remark") + 1 if "Detailed Remark" in headers else None

        label_to_col = {}
        for mapping in value_mappings:
            label = str(mapping.get("label", ""))
            col_name = str(mapping.get("f1_col" if side == "f1" else "f2_col", ""))
            target_name = label if label in headers else col_name
            if label and target_name and target_name in headers:
                label_to_col[label] = headers.index(target_name) + 1

        total_rows = max(0, ws.max_row - data_start_row + 1)
        style_fast_mode = fast_mode or (total_rows > FAST_STYLE_ROW_THRESHOLD)

        # Number/date formats: apply fully only for moderate row counts.
        if total_rows <= NUMBER_FORMAT_ROW_LIMIT and not fast_mode:
            for mapping in value_mappings:
                col_name = str(mapping.get("f1_col" if side == "f1" else "f2_col", ""))
                if not col_name or col_name not in headers:
                    continue
                col_idx = headers.index(col_name) + 1
                fmt = "DD-MMM-YYYY" if "date" in col_name.casefold() else "#,##0.00"
                for row_idx in range(data_start_row, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = fmt

        # Build maps from DataFrame once.
        status_series = df["Match Status"].tolist()
        mismatch_series = df["_mismatch_labels"].tolist()
        fast_color_cols = [c for c in (status_col, remark_col, detail_col) if c]

        if fast_mode:
            # Ultra-fast path for very large datasets: color only status cells.
            if status_col:
                for row_idx in range(data_start_row, ws.max_row + 1):
                    data_idx = row_idx - data_start_row
                    if data_idx >= len(status_series):
                        continue
                    status = str(status_series[data_idx])
                    fill = STATUS_FILL_CACHE.get(status)
                    font = STATUS_FONT_CACHE.get(status)
                    if fill is None or font is None:
                        continue
                    c = ws.cell(row=row_idx, column=status_col)
                    c.fill = fill
                    c.font = font
            ws.sheet_format.defaultRowHeight = 18
            ws.freeze_panes = f"A{data_start_row}"
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
            self._autofit_columns(ws, header_row=header_row, data_start_row=data_start_row, sample_rows=40)
            return
        
        for row_idx in range(data_start_row, ws.max_row + 1):
            data_idx = row_idx - data_start_row
            if data_idx >= len(status_series):
                continue
                
            status = str(status_series[data_idx]) if data_idx < len(status_series) else ""
            is_issue = status not in {"Match", "Matched", ""}
            
            if is_issue:
                row_fill = STATUS_FILL_CACHE.get(status)
                row_font = STATUS_FONT_CACHE.get(status)
                if row_fill is not None and row_font is not None:
                    style_cols = fast_color_cols if style_fast_mode else range(1, ws.max_column + 1)
                    for col_idx in style_cols:
                        c = ws.cell(row=row_idx, column=col_idx)
                        c.fill = row_fill
                        c.font = row_font
                        if not style_fast_mode:
                            c.border = THIN_BORDER
                            c.alignment = ALIGN_LEFT
            elif row_idx % 2 == 0:
                for col_idx in range(1, min(4, ws.max_column + 1)):
                    ws.cell(row=row_idx, column=col_idx).fill = ALT_ROW_FILL

            # Mismatch cell highlighting
            labels = mismatch_series[data_idx] if data_idx < len(mismatch_series) else set()
            if labels and (not style_fast_mode or total_rows <= MISMATCH_HIGHLIGHT_ROW_LIMIT):
                for label in labels:
                    if label in label_to_col:
                        cidx = label_to_col[label]
                        c = ws.cell(row=row_idx, column=cidx)
                        c.fill = MISMATCH_CELL_FILL
                        c.font = MISMATCH_CELL_FONT

            # Qty mismatch
            if status == "Qty Mismatch" and qty_col_idx:
                qc = ws.cell(row=row_idx, column=qty_col_idx)
                qc.fill = QTY_MISMATCH_FILL
                qc.font = QTY_MISMATCH_FONT

        ws.sheet_format.defaultRowHeight = 18
        ws.freeze_panes = f"A{data_start_row}"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
        self._autofit_columns(ws, header_row=header_row, data_start_row=data_start_row)

    def _style_mismatch_sheet(
        self,
        ws,
        mismatch_df: pd.DataFrame,
        *,
        brand_file_name: str,
        essgee_file_name: str,
        fast_mode: bool = False,
    ) -> None:
        # Summary block rows 1-8
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(6, ws.max_column))
        ws.cell(row=1, column=1).value = "BRAND MISMATCH REPORT"
        ws.cell(row=1, column=1).fill = PatternFill(fill_type="solid", fgColor="0F172A")
        ws.cell(row=1, column=1).font = Font(color="FFFFFF", bold=True, size=15)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        now = datetime.now().strftime("%d-%b-%Y %H:%M")
        ws.cell(row=2, column=1).value = f"Generated: {now}"
        ws.cell(row=3, column=1).value = f"Brand File: {brand_file_name}"
        ws.cell(row=4, column=1).value = f"EssGee File: {essgee_file_name}"

        mismatch = int(
            mismatch_df["Match Status"].isin(["Mismatch", "Brand Mismatch", "Value Mismatch"]).sum()
        ) if "Match Status" in mismatch_df.columns else 0
        qty_mismatch = int((mismatch_df["Match Status"] == "Qty Mismatch").sum()) if "Match Status" in mismatch_df.columns else 0
        not_in_data = int(
            mismatch_df["Match Status"].isin(
                ["Only In F1", "Only In F2", "Not In Data", "Not In Brand", "Not In EssGee"]
            ).sum()
        ) if "Match Status" in mismatch_df.columns else 0
        total = len(mismatch_df)
        ws.cell(row=6, column=1).value = (
            f"SUMMARY | Mismatch: {mismatch} | Qty Mismatch: {qty_mismatch} | "
            f"Not In Data: {not_in_data} | Total: {total}"
        )

        header_row = 8
        for c in ws[header_row]:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = ALIGN_CENTER

        headers = [cell.value for cell in ws[header_row]]
        status_col = headers.index("Match Status") + 1 if "Match Status" in headers else 1
        status_series = mismatch_df["Match Status"].tolist()
        total_rows = max(0, ws.max_row - 8)
        style_fast_mode = fast_mode or (total_rows > FAST_STYLE_ROW_THRESHOLD)
        fast_color_cols = [c for c in (status_col, 2, 3) if c <= ws.max_column]

        if fast_mode:
            # Ultra-fast path: color only status column.
            if status_col:
                for idx, row_idx in enumerate(range(9, ws.max_row + 1)):
                    status = str(status_series[idx]) if idx < len(status_series) else "Value Mismatch"
                    fill = STATUS_FILL_CACHE.get(status, STATUS_FILL_CACHE["Value Mismatch"])
                    font = STATUS_FONT_CACHE.get(status, STATUS_FONT_CACHE["Value Mismatch"])
                    c = ws.cell(row=row_idx, column=status_col)
                    c.fill = fill
                    c.font = font
            ws.freeze_panes = "A9"
            ws.auto_filter.ref = f"A8:{get_column_letter(ws.max_column)}{ws.max_row}"
            self._autofit_columns(ws, header_row=8, data_start_row=9, sample_rows=40)
            return

        # Apply row styling.
        for idx, row_idx in enumerate(range(9, ws.max_row + 1)):
            status = str(status_series[idx]) if idx < len(status_series) else "Value Mismatch"
            fill = STATUS_FILL_CACHE.get(status, STATUS_FILL_CACHE["Value Mismatch"])
            font = STATUS_FONT_CACHE.get(status, STATUS_FONT_CACHE["Value Mismatch"])
            style_cols = fast_color_cols if style_fast_mode else range(1, ws.max_column + 1)
            for col_idx in style_cols:
                c = ws.cell(row=row_idx, column=col_idx)
                c.fill = fill
                c.font = font

        # Totals row styling
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1).value = f"Totals: {total}"
        ws.cell(row=total_row, column=2).value = f"Mismatch={mismatch}"
        ws.cell(row=total_row, column=3).value = f"Qty Mismatch={qty_mismatch}"
        ws.cell(row=total_row, column=4).value = f"Not In Data={not_in_data}"
        for col_idx in range(1, min(6, ws.max_column + 1)):
            c = ws.cell(row=total_row, column=col_idx)
            c.fill = PatternFill(fill_type="solid", fgColor="1F2937")
            c.font = Font(color="FFFFFF", bold=True)

        ws.freeze_panes = "A9"
        ws.auto_filter.ref = f"A8:{get_column_letter(ws.max_column)}{ws.max_row}"
        self._autofit_columns(ws, header_row=8, data_start_row=9)

    def _style_summary_sheet(
        self,
        ws,
        summary_df: pd.DataFrame,
        *,
        brand_file_name: str,
        essgee_file_name: str,
        fast_mode: bool = False,
    ) -> None:
        headers = [cell.value for cell in ws[1]]
        for c in ws[1]:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER

        status_col = headers.index("Match Status") + 1 if "Match Status" in headers else 1
        remark_col = headers.index("Match Remark") + 1 if "Match Remark" in headers else None
        detail_col = headers.index("Detailed Remark") + 1 if "Detailed Remark" in headers else None
        total_rows = max(0, ws.max_row - 1)
        style_fast_mode = fast_mode or (total_rows > FAST_STYLE_ROW_THRESHOLD)

        # Build lookups from DataFrame
        mismatch_series = summary_df["_mismatch_labels"].tolist() if "_mismatch_labels" in summary_df.columns else []
        status_series = summary_df["Match Status"].tolist()

        label_to_pair_cols: dict[str, tuple[int, int]] = {}
        for idx, name in enumerate(headers, start=1):
            if not isinstance(name, str):
                continue
            if name.startswith("Brand "):
                label = name[len("Brand ") :]
                ess_name = f"EssGee {label}"
                if ess_name in headers:
                    label_to_pair_cols[label] = (idx, headers.index(ess_name) + 1)

        fast_color_cols = [c for c in (status_col, remark_col, detail_col) if c]

        if fast_mode:
            # Ultra-fast path: color only status column.
            if status_col:
                for row_idx in range(2, ws.max_row + 1):
                    data_idx = row_idx - 2
                    if data_idx >= len(status_series):
                        continue
                    status = str(status_series[data_idx])
                    fill = STATUS_FILL_CACHE.get(status)
                    font = STATUS_FONT_CACHE.get(status)
                    if fill is None or font is None:
                        continue
                    c = ws.cell(row=row_idx, column=status_col)
                    c.fill = fill
                    c.font = font
            ws.sheet_format.defaultRowHeight = 18
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            self._autofit_columns(ws, header_row=1, data_start_row=2, sample_rows=40)
            return

        # Single pass styling.
        for row_idx in range(2, ws.max_row + 1):
            data_idx = row_idx - 2
            status = str(status_series[data_idx]) if data_idx < len(status_series) else ""
            is_issue = status not in {"Match", "Matched", ""}
            
            if is_issue:
                row_fill = STATUS_FILL_CACHE.get(status)
                row_font = STATUS_FONT_CACHE.get(status)
                if row_fill is not None and row_font is not None:
                    style_cols = fast_color_cols if style_fast_mode else range(1, ws.max_column + 1)
                    for col_idx in style_cols:
                        c = ws.cell(row=row_idx, column=col_idx)
                        c.fill = row_fill
                        c.font = row_font
                        if not style_fast_mode:
                            c.border = THIN_BORDER
                            c.alignment = ALIGN_LEFT
            elif row_idx % 2 == 0:
                for col_idx in range(1, min(4, ws.max_column + 1)):
                    ws.cell(row=row_idx, column=col_idx).fill = ALT_ROW_FILL

            # Mismatch highlighting
            labels = mismatch_series[data_idx] if data_idx < len(mismatch_series) else set()
            if labels and (not style_fast_mode or total_rows <= MISMATCH_HIGHLIGHT_ROW_LIMIT):
                for label in labels:
                    pair = label_to_pair_cols.get(label)
                    if not pair:
                        continue
                    for cidx in pair:
                        c = ws.cell(row=row_idx, column=cidx)
                        c.fill = MISMATCH_CELL_FILL
                        c.font = MISMATCH_CELL_FONT

        ws.sheet_format.defaultRowHeight = 18
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        self._autofit_columns(ws, header_row=1, data_start_row=2)

    def _autofit_columns(
        self,
        ws,
        *,
        header_row: int = 1,
        data_start_row: int = 2,
        sample_rows: int = 100,
    ) -> None:
        if ws.max_column <= 0:
            return

        max_row = ws.max_row
        sample_end = min(max_row, max(data_start_row, data_start_row + sample_rows - 1))
        
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            header_val = ws.cell(row=header_row, column=col_idx).value
            if header_val is not None:
                max_len = len(str(header_val))

            for row_idx in range(data_start_row, sample_end + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                value_len = len(str(value))
                if value_len > max_len:
                    max_len = value_len

            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(10, min(60, int(max_len * 1.1) + 2))
