from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd


EXCEL_XML_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
OPEN_DOCUMENT_EXTENSIONS = {".ods", ".odf", ".odt"}
DELIMITED_EXTENSIONS = {".csv", ".tsv"}
SUPPORTED_EXTENSIONS = EXCEL_XML_EXTENSIONS | OPEN_DOCUMENT_EXTENSIONS | DELIMITED_EXTENSIONS | {".xls", ".xlsb"}
SUPPORTED_EXTENSIONS_TEXT = ", ".join(sorted(SUPPORTED_EXTENSIONS))


class LoaderError(Exception):
    """Raised when file loading fails with a user-friendly message."""


def _validate_path(path: str | Path) -> Path:
    file_path = Path(path).expanduser().resolve()
    if not file_path.exists():
        raise LoaderError(f"File not found: {file_path}")
    if not file_path.is_file():
        raise LoaderError(f"Path is not a file: {file_path}")
    if file_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise LoaderError(
            f"Unsupported file type '{file_path.suffix}'. Supported: {SUPPORTED_EXTENSIONS_TEXT}"
        )
    return file_path


def _normalize_header_row(header_row: int | None) -> int | None:
    if header_row is None:
        return 0
    if not isinstance(header_row, int):
        raise LoaderError("header_row must be an integer or null.")
    if header_row < 0:
        raise LoaderError("header_row must be >= 0.")
    return header_row


def _read_delimited(file_path: Path, header_row: int | None, nrows: int | None = None) -> pd.DataFrame:
    sep = "\t" if file_path.suffix.lower() == ".tsv" else ","
    try:
        return pd.read_csv(
            file_path,
            sep=sep,
            dtype=object,
            header=header_row,
            nrows=nrows,
            low_memory=False,
            encoding_errors="ignore",
        )
    except UnicodeDecodeError:
        return pd.read_csv(
            file_path,
            sep=sep,
            dtype=object,
            header=header_row,
            nrows=nrows,
            low_memory=False,
            encoding="latin-1",
        )
    except Exception as exc:
        raise LoaderError(f"Unable to read delimited file: {exc}") from exc


def _excel_reader_settings(suffix: str) -> tuple[str | None, dict[str, Any] | None]:
    if suffix in EXCEL_XML_EXTENSIONS:
        # Read-only mode is significantly faster/lighter for large workbooks.
        return "openpyxl", {"read_only": True, "data_only": True}
    if suffix == ".xls":
        return "xlrd", None
    if suffix == ".xlsb":
        return "pyxlsb", None
    if suffix in OPEN_DOCUMENT_EXTENSIONS:
        return "odf", None
    return None, None


def _read_excel(
    file_path: Path, sheet: str | int | None, header_row: int | None, nrows: int | None = None
) -> pd.DataFrame:
    suffix = file_path.suffix.lower()
    engine, engine_kwargs = _excel_reader_settings(suffix)

    try:
        return pd.read_excel(
            file_path,
            sheet_name=sheet if sheet is not None else 0,
            header=header_row,
            dtype=object,
            nrows=nrows,
            engine=engine,
            engine_kwargs=engine_kwargs,
        )
    except ImportError as exc:
        if suffix == ".xls":
            raise LoaderError(
                "Reading .xls requires the optional dependency 'xlrd'. "
                "Install it with: pip install xlrd"
            ) from exc
        if suffix == ".xlsb":
            raise LoaderError(
                "Reading .xlsb requires the optional dependency 'pyxlsb'. "
                "Install it with: pip install pyxlsb"
            ) from exc
        if suffix in OPEN_DOCUMENT_EXTENSIONS:
            raise LoaderError(
                "Reading .ods/.odf/.odt requires the optional dependency 'odfpy'. "
                "Install it with: pip install odfpy"
            ) from exc
        raise LoaderError(f"Missing dependency for Excel reading: {exc}") from exc
    except ValueError as exc:
        raise LoaderError(f"Invalid sheet selection '{sheet}': {exc}") from exc
    except Exception as exc:
        raise LoaderError(f"Unable to read Excel file: {exc}") from exc


def _clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    columns = []
    for idx, col in enumerate(df.columns):
        if col is None or (isinstance(col, float) and np.isnan(col)):
            columns.append(f"Unnamed_{idx+1}")
            continue
        col_text = str(col).strip()
        columns.append(col_text if col_text else f"Unnamed_{idx+1}")
    df.columns = columns
    return df


def load_excel(
    path: str | Path,
    sheet: str | int | None = None,
    header_row: int | None = 0,
    nrows: int | None = None,
) -> pd.DataFrame:
    """
    Load Excel/OpenDocument/delimited files into a pandas DataFrame.
    Uses dtype=object to preserve identifiers exactly (invoice/barcode values).
    """
    file_path = _validate_path(path)
    normalized_header = _normalize_header_row(header_row)

    if file_path.suffix.lower() in DELIMITED_EXTENSIONS:
        df = _read_delimited(file_path, normalized_header, nrows=nrows)
    else:
        df = _read_excel(file_path, sheet=sheet, header_row=normalized_header, nrows=nrows)

    if isinstance(df, dict):
        raise LoaderError("Multiple sheets returned unexpectedly. Choose a single sheet.")

    return _clean_columns(df)


def fast_row_count(path: str | Path, sheet: str | int | None = None, header_row: int | None = 0) -> int:
    """
    Fast row count estimate for large files.
    Returns data-row count (excluding header row offset).
    """
    file_path = _validate_path(path)
    normalized_header = _normalize_header_row(header_row) or 0
    header_offset = normalized_header + 1
    suffix = file_path.suffix.lower()

    if suffix in DELIMITED_EXTENSIONS:
        total_lines = 0
        with file_path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle)
            for _ in reader:
                total_lines += 1
        return max(total_lines - header_offset, 0)

    if suffix in EXCEL_XML_EXTENSIONS:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise LoaderError(f"openpyxl unavailable for row counting: {exc}") from exc

        wb = load_workbook(file_path, read_only=True, data_only=True, keep_links=False)
        try:
            if sheet is None:
                ws = wb.worksheets[0]
            elif isinstance(sheet, int):
                ws = wb.worksheets[sheet]
            else:
                ws = wb[str(sheet)]
            total_rows = int(ws.max_row or 0)
        finally:
            wb.close()
        return max(total_rows - header_offset, 0)

    if suffix == ".xls":
        try:
            import xlrd
        except Exception as exc:
            raise LoaderError(f"xlrd unavailable for .xls row counting: {exc}") from exc

        wb = xlrd.open_workbook(file_path, on_demand=True)
        try:
            if sheet is None:
                ws = wb.sheet_by_index(0)
            elif isinstance(sheet, int):
                ws = wb.sheet_by_index(sheet)
            else:
                ws = wb.sheet_by_name(str(sheet))
            total_rows = int(ws.nrows or 0)
        finally:
            try:
                wb.release_resources()
            except Exception:
                pass
        return max(total_rows - header_offset, 0)

    if suffix == ".xlsb":
        try:
            from pyxlsb import open_workbook
        except Exception as exc:
            raise LoaderError(f"pyxlsb unavailable for .xlsb row counting: {exc}") from exc

        try:
            with open_workbook(str(file_path)) as wb:
                sheet_names = [str(name) for name in wb.sheets]
                sheet_name = _resolve_sheet_name(sheet_names, sheet)
                with wb.get_sheet(sheet_name) as ws:
                    total_rows = sum(1 for _ in ws.rows())
        except Exception as exc:
            raise LoaderError(f"Unable to count .xlsb rows: {exc}") from exc
        return max(int(total_rows) - header_offset, 0)

    if suffix in OPEN_DOCUMENT_EXTENSIONS:
        df = _read_excel(file_path, sheet=sheet, header_row=None, nrows=None)
        return max(len(df.index) - header_offset, 0)

    raise LoaderError(f"Unsupported file type for row counting: {suffix}")


def list_sheets(path: str | Path) -> list[str]:
    """Return available sheet names. CSV returns a single virtual sheet."""
    file_path = _validate_path(path)
    if file_path.suffix.lower() in DELIMITED_EXTENSIONS:
        return ["Delimited Data"]

    suffix = file_path.suffix.lower()
    engine, engine_kwargs = _excel_reader_settings(suffix)
    try:
        with pd.ExcelFile(file_path, engine=engine, engine_kwargs=engine_kwargs) as workbook:
            return [str(name) for name in workbook.sheet_names]
    except ImportError as exc:
        if suffix == ".xls":
            raise LoaderError(
                "Reading .xls requires the optional dependency 'xlrd'. "
                "Install it with: pip install xlrd"
            ) from exc
        if suffix == ".xlsb":
            raise LoaderError(
                "Reading .xlsb requires the optional dependency 'pyxlsb'. "
                "Install it with: pip install pyxlsb"
            ) from exc
        if suffix in OPEN_DOCUMENT_EXTENSIONS:
            raise LoaderError(
                "Reading .ods/.odf/.odt requires the optional dependency 'odfpy'. "
                "Install it with: pip install odfpy"
            ) from exc
        raise LoaderError(f"Missing dependency for sheet listing: {exc}") from exc
    except Exception as exc:
        raise LoaderError(f"Unable to list sheets: {exc}") from exc

    raise LoaderError(f"Unable to list sheets for unsupported type: {suffix}")


def _resolve_sheet_name(sheet_names: list[str], sheet: str | int | None) -> str:
    if not sheet_names:
        raise LoaderError("Workbook has no sheets.")
    if sheet is None:
        return sheet_names[0]
    if isinstance(sheet, int):
        try:
            return sheet_names[sheet]
        except IndexError as exc:
            raise LoaderError(f"Invalid sheet index '{sheet}'.") from exc
    return str(sheet)


def _to_json_safe(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (pd.Timestamp, np.datetime64)):
        return pd.to_datetime(value, errors="coerce").isoformat() if pd.notna(value) else None
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        if pd.isna(value) or np.isinf(value):
            return None
        return float(value)
    if pd.isna(value):
        return None
    return value


def preview_rows(df: pd.DataFrame, n: int = 5) -> list[dict[str, Any]]:
    """Return first n rows as JSON-safe list of dict records."""
    if n <= 0:
        return []

    preview = df.head(n).replace({np.nan: None})
    records: list[dict[str, Any]] = []
    for _, row in preview.iterrows():
        record = {str(col): _to_json_safe(val) for col, val in row.items()}
        records.append(record)
    return records
