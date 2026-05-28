#!/usr/bin/env python
"""Quick smoke test for backend startup and split-qty matching."""

from pathlib import Path
import sys
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent / "backend"))

from core.exporter import ExcelExporter
from core.loader import SUPPORTED_EXTENSIONS, fast_row_count, list_sheets, load_excel
from core.matcher import MatchConfig, MatchingEngine
from core.normalizer import normalize_bill_number
from core.remarks_engine import apply_remarks
from server import create_app


def build_config() -> MatchConfig:
    return MatchConfig.from_dict(
        {
            "key_columns": [
                {
                    "f1_col": "Invoice No",
                    "f2_col": "Invoice No",
                    "label": "Invoice No",
                    "match_type": "bill",
                },
                {
                    "f1_col": "Party Name",
                    "f2_col": "Party Name",
                    "label": "Party Name",
                    "match_type": "identifier",
                },
            ],
            "value_columns": [
                {
                    "f1_col": "Qty",
                    "f2_col": "Qty",
                    "label": "Qty",
                    "match_type": "number",
                    "tolerance": 0,
                },
                {
                    "f1_col": "MRP",
                    "f2_col": "MRP",
                    "label": "MRP",
                    "match_type": "number",
                    "tolerance": 0,
                },
            ],
            "fuzzy_enabled": True,
            "qty_expansion_enabled": False,
            "case_insensitive": True,
            "trim": True,
        }
    )


def run_split_smoke_test() -> None:
    f1 = pd.DataFrame(
        [
            {"Invoice No": "INV1", "Party Name": "ABC", "Qty": 2, "MRP": 100},
        ]
    )
    f2 = pd.DataFrame(
        [
            {"Invoice No": "INV1", "Party Name": "ABC", "Qty": 1, "MRP": 50},
            {"Invoice No": "INV1", "Party Name": "ABC", "Qty": 1, "MRP": 50},
        ]
    )
    reverse_f1 = pd.DataFrame(
        [
            {"Invoice No": "INV2", "Party Name": "ABC", "Qty": 1, "MRP": 50},
            {"Invoice No": "INV2", "Party Name": "ABC", "Qty": 1, "MRP": 50},
        ]
    )
    reverse_f2 = pd.DataFrame(
        [
            {"Invoice No": "INV2", "Party Name": "ABC", "Qty": 2, "MRP": 100},
        ]
    )
    both_split_f1 = pd.DataFrame(
        [
            {"Invoice No": "INV3", "Party Name": "ABC", "Qty": 1, "MRP": 50},
            {"Invoice No": "INV3", "Party Name": "ABC", "Qty": 1, "MRP": 50},
        ]
    )
    both_split_f2 = pd.DataFrame(
        [
            {"Invoice No": "INV3", "Party Name": "ABC", "Qty": 1, "MRP": 50},
            {"Invoice No": "INV3", "Party Name": "ABC", "Qty": 1, "MRP": 50},
        ]
    )

    engine = MatchingEngine(build_config())
    forward_rows = apply_remarks(engine.run(f1, f2)["rows"])
    forward_statuses = [row["match_status"] for row in forward_rows]
    if not all(status == "Match" for status in forward_statuses):
        raise RuntimeError(f"Forward split quantity case failed: {forward_statuses}")

    reverse_rows = apply_remarks(MatchingEngine(build_config()).run(reverse_f1, reverse_f2)["rows"])
    reverse_statuses = [row["match_status"] for row in reverse_rows]
    if not all(status == "Match" for status in reverse_statuses):
        raise RuntimeError(f"Reverse split quantity case failed: {reverse_statuses}")

    both_split_rows = apply_remarks(MatchingEngine(build_config()).run(both_split_f1, both_split_f2)["rows"])
    both_split_statuses = [row["match_status"] for row in both_split_rows]
    if not all(status == "Match" for status in both_split_statuses):
        raise RuntimeError(f"Both-sides split quantity case failed: {both_split_statuses}")

    numeric_key_engine = MatchingEngine(build_config())
    brand_single_ok = numeric_key_engine._brand_single_split_values_match(
        {"Invoice No": 1001, "Party Name": "ABC", "Qty": 2, "MRP": 100},
        [
            {"Invoice No": 1001, "Party Name": "ABC", "Qty": 1, "MRP": 50},
            {"Invoice No": 1001, "Party Name": "ABC", "Qty": 1, "MRP": 50},
        ],
    )
    if not brand_single_ok:
        raise RuntimeError("Numeric invoice split key comparison failed")

    summarized_key_ok = numeric_key_engine._split_group_values_match(
        {
            0: {"Invoice No": 1002, "Party Name": "ABC", "Qty": 1, "MRP": 50},
            1: {"Invoice No": 1002, "Party Name": "ABC", "Qty": 1, "MRP": 50},
        },
        {0: {"Invoice No": 1002, "Party Name": "ABC", "Qty": 2, "MRP": 100}},
    )
    if not summarized_key_ok:
        raise RuntimeError("F1 split to F2 summarized key comparison failed")

    print("OK Forward split quantity case matched")
    print("OK Reverse split quantity case matched")
    print("OK Both-sides split quantity case matched")
    print("OK Numeric invoice split key comparison matched")


def run_invoice_padding_smoke_test() -> None:
    if normalize_bill_number("1AS-0018922 ") != normalize_bill_number("01AS-0018922   "):
        raise RuntimeError("Leading-zero invoice normalization failed")

    config = MatchConfig.from_dict(
        {
            "key_columns": [
                {
                    "f1_col": "Bill No.",
                    "f2_col": "Bill No.",
                    "label": "Bill No.",
                    "match_type": "bill",
                },
                {
                    "f1_col": "Party Name",
                    "f2_col": "Party Name",
                    "label": "Party Name",
                    "match_type": "identifier",
                },
            ],
            "value_columns": [
                {
                    "f1_col": "Qty",
                    "f2_col": "Qty",
                    "label": "Qty",
                    "match_type": "number",
                    "tolerance": 0,
                },
                {
                    "f1_col": "Amount",
                    "f2_col": "Amount",
                    "label": "Amount",
                    "match_type": "number",
                    "tolerance": 0,
                },
            ],
            "fuzzy_enabled": True,
            "fuzzy_threshold": 85,
            "qty_expansion_enabled": False,
            "case_insensitive": True,
            "trim": True,
        }
    )
    f1 = pd.DataFrame(
        [
            {"Bill No.": "1AS-0018922 ", "Party Name": "ABC", "Qty": 2, "Amount": 100},
        ]
    )
    f2 = pd.DataFrame(
        [
            {"Bill No.": "01AS-0018922   ", "Party Name": "ABC", "Qty": 2, "Amount": 100},
        ]
    )
    rows = apply_remarks(MatchingEngine(config).run(f1, f2)["rows"])
    statuses = [row["match_status"] for row in rows]
    if not all(status == "Match" for status in statuses):
        raise RuntimeError(f"Leading-zero invoice match failed: {statuses}")

    print("OK Leading-zero invoice format matched")


def run_qty_shortage_remark_smoke_test() -> None:
    f1 = pd.DataFrame(
        [
            {"Invoice No": "INV4", "Party Name": "ABC", "Qty": 1, "MRP": 50},
            {"Invoice No": "INV4", "Party Name": "ABC", "Qty": 1, "MRP": 50},
        ]
    )
    f2 = pd.DataFrame(
        [
            {"Invoice No": "INV4", "Party Name": "ABC", "Qty": 1, "MRP": 50},
        ]
    )
    rows = apply_remarks(MatchingEngine(build_config()).run(f1, f2)["rows"])
    qty_rows = [row for row in rows if row["match_status"] == "Qty Mismatch"]
    if len(qty_rows) != len(rows):
        raise RuntimeError(f"Brand extra qty shortage status failed: {[row['match_status'] for row in rows]}")
    remarks = sorted(str(row.get("match_remark", "")) for row in qty_rows)
    if remarks != ["All Match", "Not In"]:
        raise RuntimeError(f"Brand extra qty shortage remark failed: {[row.get('match_remark') for row in rows]}")
    all_match_rows = [row for row in qty_rows if row.get("match_remark") == "All Match"]
    if len(all_match_rows) != 1 or all_match_rows[0].get("detailed_remark") != "All values matched":
        raise RuntimeError(f"Brand matched qty detail failed: {[row.get('detailed_remark') for row in rows]}")
    not_in_rows = [row for row in qty_rows if row.get("match_remark") == "Not In"]
    not_in_detail = str(not_in_rows[0].get("detailed_remark", "")) if len(not_in_rows) == 1 else ""
    if "Brand File Qty = 2" not in not_in_detail or "EssGee File Qty = 1" not in not_in_detail:
        raise RuntimeError(f"Brand extra qty shortage detail failed: {[row.get('detailed_remark') for row in rows]}")

    print("OK Brand extra qty shortage remark matched")


def run_key_mismatch_not_qty_smoke_test() -> None:
    different_invoice_f1 = pd.DataFrame(
        [
            {"Invoice No": "INV-001", "Party Name": "ABC TRADERS", "Qty": 2, "MRP": 50},
        ]
    )
    different_invoice_f2 = pd.DataFrame(
        [
            {"Invoice No": "INV-002", "Party Name": "ABC TRADEES", "Qty": 1, "MRP": 50},
        ]
    )
    different_invoice_rows = apply_remarks(
        MatchingEngine(build_config()).run(different_invoice_f1, different_invoice_f2)["rows"]
    )
    different_invoice_statuses = [row["match_status"] for row in different_invoice_rows]
    if "Qty Mismatch" in different_invoice_statuses:
        raise RuntimeError(f"Different invoice was incorrectly marked Qty Mismatch: {different_invoice_statuses}")
    if set(different_invoice_statuses) != {"Not In EssGee", "Not In Brand"}:
        raise RuntimeError(f"Different invoice not-in remark failed: {different_invoice_statuses}")

    different_party_f1 = pd.DataFrame(
        [
            {"Invoice No": "INV-005", "Party Name": "ABC TRADERS", "Qty": 2, "MRP": 50},
        ]
    )
    different_party_f2 = pd.DataFrame(
        [
            {"Invoice No": "INV-005", "Party Name": "ABCD TRADERS", "Qty": 1, "MRP": 50},
        ]
    )
    different_party_rows = apply_remarks(
        MatchingEngine(build_config()).run(different_party_f1, different_party_f2)["rows"]
    )
    different_party_statuses = [row["match_status"] for row in different_party_rows]
    if different_party_statuses != ["Value Mismatch"]:
        raise RuntimeError(f"Different party was incorrectly marked: {different_party_statuses}")
    detail = str(different_party_rows[0].get("detailed_remark", ""))
    if "Party Name is different" not in detail or "Qty is different" not in detail:
        raise RuntimeError(f"Different party detail failed: {detail}")

    print("OK Key mismatches do not become qty mismatch")


def run_blank_barcode_split_value_smoke_test() -> None:
    config = MatchConfig.from_dict(
        {
            "key_columns": [
                {
                    "f1_col": "Invoice No",
                    "f2_col": "Invoice No",
                    "label": "Invoice No",
                    "match_type": "bill",
                },
                {
                    "f1_col": "Party Name",
                    "f2_col": "Party Name",
                    "label": "Party Name",
                    "match_type": "identifier",
                },
                {
                    "f1_col": "EAN",
                    "f2_col": "EAN",
                    "label": "EAN",
                    "match_type": "barcode",
                },
            ],
            "value_columns": [
                {
                    "f1_col": "Qty",
                    "f2_col": "Qty",
                    "label": "Qty",
                    "match_type": "number",
                    "tolerance": 0,
                },
                {
                    "f1_col": "MRP Value",
                    "f2_col": "MRP Value",
                    "label": "MRP Value",
                    "match_type": "number",
                    "tolerance": 0,
                },
                {
                    "f1_col": "Discount Value",
                    "f2_col": "Discount Value",
                    "label": "Discount Value",
                    "match_type": "number",
                    "tolerance": 0,
                },
                {
                    "f1_col": "Net Value",
                    "f2_col": "Net Value",
                    "label": "Net Value",
                    "match_type": "number",
                    "tolerance": 0,
                },
            ],
            "fuzzy_enabled": True,
            "qty_expansion_enabled": False,
            "case_insensitive": True,
            "trim": True,
        }
    )
    f1 = pd.DataFrame(
        [
            {"Invoice No": "INV8", "Party Name": "ABC", "EAN": "", "Qty": 1, "MRP Value": 1350, "Discount Value": 0, "Net Value": 1350},
            {"Invoice No": "INV8", "Party Name": "ABC", "EAN": "", "Qty": 1, "MRP Value": 2090, "Discount Value": 0, "Net Value": 2090},
            {"Invoice No": "INV8", "Party Name": "ABC", "EAN": "", "Qty": 1, "MRP Value": 2090, "Discount Value": 0, "Net Value": 2090},
        ]
    )
    f2 = pd.DataFrame(
        [
            {"Invoice No": "INV8", "Party Name": "ABC", "EAN": "8901", "Qty": 1, "MRP Value": 1350, "Discount Value": 0, "Net Value": 1350},
            {"Invoice No": "INV8", "Party Name": "ABC", "EAN": "8902", "Qty": 2, "MRP Value": 4180, "Discount Value": 0, "Net Value": 4180},
        ]
    )
    rows = apply_remarks(MatchingEngine(config).run(f1, f2)["rows"])
    statuses = [row["match_status"] for row in rows]
    if not all(status == "Match" for status in statuses):
        raise RuntimeError(f"Blank barcode split/value total match failed: {statuses}")

    conflict_f1 = pd.DataFrame(
        [{"Invoice No": "INV9", "Party Name": "ABC", "EAN": "111", "Qty": 1, "MRP Value": 100, "Discount Value": 0, "Net Value": 100}]
    )
    conflict_f2 = pd.DataFrame(
        [{"Invoice No": "INV9", "Party Name": "ABC", "EAN": "222", "Qty": 1, "MRP Value": 100, "Discount Value": 0, "Net Value": 100}]
    )
    conflict_rows = apply_remarks(MatchingEngine(config).run(conflict_f1, conflict_f2)["rows"])
    conflict_status = conflict_rows[0].get("match_status")
    conflict_remark = conflict_rows[0].get("match_remark")
    if conflict_status != "Value Mismatch" or conflict_remark != "Barcode Mismatch":
        raise RuntimeError(f"Two-sided barcode mismatch was hidden: {conflict_status}/{conflict_remark}")

    print("OK Blank barcode split totals and real barcode conflicts handled")


def run_qty_mismatch_value_detail_smoke_test() -> None:
    config = MatchConfig.from_dict(
        {
            "key_columns": [
                {
                    "f1_col": "Invoice No",
                    "f2_col": "Invoice No",
                    "label": "Invoice No",
                    "match_type": "bill",
                },
                {
                    "f1_col": "Party Name",
                    "f2_col": "Party Name",
                    "label": "Party Name",
                    "match_type": "identifier",
                },
            ],
            "value_columns": [
                {
                    "f1_col": "Qty",
                    "f2_col": "Qty",
                    "label": "Qty",
                    "match_type": "number",
                    "tolerance": 0,
                },
                {
                    "f1_col": "MRP",
                    "f2_col": "MRP",
                    "label": "MRP",
                    "match_type": "number",
                    "tolerance": 0,
                },
                {
                    "f1_col": "Discount Value",
                    "f2_col": "Discount Value",
                    "label": "Discount Value",
                    "match_type": "number",
                    "tolerance": 0,
                },
                {
                    "f1_col": "Net",
                    "f2_col": "Net",
                    "label": "Net",
                    "match_type": "number",
                    "tolerance": 0,
                },
            ],
            "fuzzy_enabled": True,
            "qty_expansion_enabled": False,
            "case_insensitive": True,
            "trim": True,
        }
    )
    f1 = pd.DataFrame(
        [
            {"Invoice No": "INV6", "Party Name": "ABC", "Qty": 2, "MRP": 100, "Discount Value": 1349, "Net": 3146},
            {"Invoice No": "INV6", "Party Name": "ABC", "Qty": 1, "MRP": 100, "Discount Value": 449, "Net": 1049},
        ]
    )
    f2 = pd.DataFrame(
        [
            {"Invoice No": "INV6", "Party Name": "ABC", "Qty": 2, "MRP": 100, "Discount Value": 1798, "Net": 2697},
        ]
    )
    rows = apply_remarks(MatchingEngine(config).run(f1, f2)["rows"])
    paired_rows = [row for row in rows if row.get("f1_index") is not None and row.get("f2_index") is not None]
    if len(paired_rows) != 1:
        raise RuntimeError(f"Qty/value mismatch paired row count failed: {len(paired_rows)}")

    paired = paired_rows[0]
    paired_detail = str(paired.get("detailed_remark", ""))
    if paired.get("match_status") != "Value Mismatch" or paired.get("match_remark") != "Discount Mismatch":
        raise RuntimeError(f"Qty/value mismatch status failed: {paired.get('match_status')}/{paired.get('match_remark')}")
    if "Qty is different" in paired_detail or "Brand File Qty" in paired_detail:
        raise RuntimeError(f"Qty/value mismatch detail incorrectly included qty totals: {paired_detail}")
    if "Discount Value is different" not in paired_detail:
        raise RuntimeError(f"Qty/value mismatch detail missed discount column: {paired_detail}")
    if "Net is different" in paired_detail:
        raise RuntimeError(f"Qty/value mismatch detail incorrectly included downstream Net column: {paired_detail}")

    missing_rows = [row for row in rows if row.get("f2_index") is None]
    missing_detail = str(missing_rows[0].get("detailed_remark", "")) if len(missing_rows) == 1 else ""
    if "Brand File Qty = 3" not in missing_detail or "EssGee File Qty = 2" not in missing_detail:
        raise RuntimeError(f"Qty shortage detail failed: {missing_detail}")

    pure_value_f1 = pd.DataFrame(
        [
            {"Invoice No": "INV7", "Party Name": "ABC", "Qty": 1, "MRP": 100, "Discount Value": 100, "Net": 900},
        ]
    )
    pure_value_f2 = pd.DataFrame(
        [
            {"Invoice No": "INV7", "Party Name": "ABC", "Qty": 1, "MRP": 100, "Discount Value": 150, "Net": 850},
        ]
    )
    pure_value_rows = apply_remarks(MatchingEngine(config).run(pure_value_f1, pure_value_f2)["rows"])
    if len(pure_value_rows) != 1:
        raise RuntimeError(f"Pure discount mismatch row count failed: {len(pure_value_rows)}")
    pure_detail = str(pure_value_rows[0].get("detailed_remark", ""))
    if pure_value_rows[0].get("match_status") != "Value Mismatch":
        raise RuntimeError(f"Pure discount mismatch status failed: {pure_value_rows[0].get('match_status')}")
    if pure_value_rows[0].get("match_remark") != "Discount Mismatch":
        raise RuntimeError(f"Pure discount mismatch remark failed: {pure_value_rows[0].get('match_remark')}")
    if "Discount Value is different" not in pure_detail or "Net is different" in pure_detail:
        raise RuntimeError(f"Pure discount mismatch detail failed: {pure_detail}")

    mrp_first_f1 = pd.DataFrame(
        [
            {"Invoice No": "INV8", "Party Name": "ABC", "Qty": 1, "MRP": 100, "Discount Value": 100, "Net": 900},
        ]
    )
    mrp_first_f2 = pd.DataFrame(
        [
            {"Invoice No": "INV8", "Party Name": "ABC", "Qty": 1, "MRP": 120, "Discount Value": 150, "Net": 850},
        ]
    )
    mrp_first_rows = apply_remarks(MatchingEngine(config).run(mrp_first_f1, mrp_first_f2)["rows"])
    if len(mrp_first_rows) != 1:
        raise RuntimeError(f"MRP-first mismatch row count failed: {len(mrp_first_rows)}")
    mrp_detail = str(mrp_first_rows[0].get("detailed_remark", ""))
    if mrp_first_rows[0].get("match_status") != "Value Mismatch":
        raise RuntimeError(f"MRP-first mismatch status failed: {mrp_first_rows[0].get('match_status')}")
    if mrp_first_rows[0].get("match_remark") != "MRP Mismatch":
        raise RuntimeError(f"MRP-first mismatch remark failed: {mrp_first_rows[0].get('match_remark')}")
    if "MRP is different" not in mrp_detail:
        raise RuntimeError(f"MRP-first mismatch detail missed MRP: {mrp_detail}")
    if "Discount Value is different" not in mrp_detail or "Net is different" not in mrp_detail:
        raise RuntimeError(f"MRP-first mismatch detail missed downstream values: {mrp_detail}")

    net_first_f1 = pd.DataFrame(
        [
            {"Invoice No": "INV9", "Party Name": "ABC", "Qty": 1, "MRP": 100, "Discount Value": 100, "Net": 900},
        ]
    )
    net_first_f2 = pd.DataFrame(
        [
            {"Invoice No": "INV9", "Party Name": "ABC", "Qty": 1, "MRP": 100, "Discount Value": 100, "Net": 850},
        ]
    )
    net_first_rows = apply_remarks(MatchingEngine(config).run(net_first_f1, net_first_f2)["rows"])
    if len(net_first_rows) != 1:
        raise RuntimeError(f"Net-first mismatch row count failed: {len(net_first_rows)}")
    net_detail = str(net_first_rows[0].get("detailed_remark", ""))
    if net_first_rows[0].get("match_status") != "Value Mismatch":
        raise RuntimeError(f"Net-first mismatch status failed: {net_first_rows[0].get('match_status')}")
    if net_first_rows[0].get("match_remark") != "Net Mismatch":
        raise RuntimeError(f"Net-first mismatch remark failed: {net_first_rows[0].get('match_remark')}")
    if "Net is different" not in net_detail:
        raise RuntimeError(f"Net-first mismatch detail missed Net: {net_detail}")

    print("OK First value mismatch remark with correct details")


def run_summary_export_smoke_test() -> None:
    key_mappings = [
        {"f1_col": "Invoice No", "f2_col": "Invoice No", "label": "Invoice No", "match_type": "bill"},
        {"f1_col": "Party Name", "f2_col": "Party Name", "label": "Party Name", "match_type": "identifier"},
    ]
    value_mappings = [
        {"f1_col": "Qty", "f2_col": "Qty", "label": "Qty", "match_type": "number", "tolerance": 0},
        {"f1_col": "MRP", "f2_col": "MRP", "label": "MRP", "match_type": "number", "tolerance": 0},
    ]
    config = MatchConfig.from_dict(
        {
            "key_columns": key_mappings,
            "value_columns": value_mappings,
            "fuzzy_enabled": True,
            "qty_expansion_enabled": False,
            "case_insensitive": True,
            "trim": True,
        }
    )
    brand_df = pd.DataFrame(
        [
            {"Invoice No": "INV9", "Party Name": "ABC", "Qty": 1, "MRP": 100, "Unused Brand": "DROP"},
            {"Invoice No": "INV10", "Party Name": "DEF", "Qty": 1, "MRP": 200, "Unused Brand": "DROP"},
        ]
    )
    essgee_df = pd.DataFrame(
        [
            {"Invoice No": "INV9", "Party Name": "ABC", "Qty": 1, "MRP": 100, "Unused EssGee": "DROP"},
            {"Invoice No": "INV10", "Party Name": "DEF", "Qty": 1, "MRP": 250, "Unused EssGee": "DROP"},
        ]
    )
    rows = apply_remarks(MatchingEngine(config).run(brand_df, essgee_df)["rows"])

    with TemporaryDirectory(prefix="excelmatcher-summary-export-") as tmpdir:
        out_path = Path(tmpdir) / "summary.xlsx"
        ExcelExporter().export_summary(
            out_path,
            brand_df=brand_df,
            essgee_df=essgee_df,
            rows=rows,
            key_mappings=key_mappings,
            value_mappings=value_mappings,
            brand_file_name="brand.xlsx",
            essgee_file_name="essgee.xlsx",
        )
        wb = load_workbook(out_path, read_only=True)
        try:
            expected_sheets = ["Brand Data", "Ess Gee Data", "Brand Mismatch Data"]
            if wb.sheetnames != expected_sheets:
                raise RuntimeError(f"Summary export sheet names failed: {wb.sheetnames}")

            brand_headers = [cell.value for cell in wb["Brand Data"][2]]
            essgee_headers = [cell.value for cell in wb["Ess Gee Data"][2]]
            mismatch_headers = [cell.value for cell in wb["Brand Mismatch Data"][8]]
            expected_headers = ["Invoice No", "Party Name", "MRP", "Qty", "Match Status", "Match Remark", "Detailed Remark"]
            if brand_headers != expected_headers:
                raise RuntimeError(f"Brand summary headers failed: {brand_headers}")
            if essgee_headers != expected_headers:
                raise RuntimeError(f"EssGee summary headers failed: {essgee_headers}")
            if mismatch_headers != expected_headers:
                raise RuntimeError(f"Mismatch summary headers failed: {mismatch_headers}")
            if any("Unused" in str(header or "") for header in brand_headers + essgee_headers + mismatch_headers):
                raise RuntimeError("Summary export included unconfigured columns")
            if wb["Brand Mismatch Data"].max_row < 9:
                raise RuntimeError("Summary export mismatch sheet has no data rows")
        finally:
            wb.close()

    print("OK Summary export uses configured columns only")


def run_loader_format_support_smoke_test() -> None:
    expected = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls", ".xlsb", ".ods", ".odf", ".odt", ".csv", ".tsv"}
    missing = expected - SUPPORTED_EXTENSIONS
    if missing:
        raise RuntimeError(f"Supported extension list missing: {sorted(missing)}")

    data = pd.DataFrame(
        [
            {"Invoice No": "INV1", "Party Name": "ABC", "Qty": 1},
            {"Invoice No": "INV2", "Party Name": "DEF", "Qty": 2},
        ]
    )
    with TemporaryDirectory(prefix="excelmatcher-loader-formats-") as tmpdir:
        tmp = Path(tmpdir)
        base_xlsx = tmp / "sample.xlsx"
        data.to_excel(base_xlsx, index=False)
        xlsx_bytes = base_xlsx.read_bytes()

        for ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            path = base_xlsx if ext == ".xlsx" else tmp / f"sample{ext}"
            if ext != ".xlsx":
                path.write_bytes(xlsx_bytes)
            if list_sheets(path) != ["Sheet1"]:
                raise RuntimeError(f"Sheet listing failed for {ext}: {list_sheets(path)}")
            loaded = load_excel(path)
            if loaded.shape[0] != 2 or "Invoice No" not in loaded.columns:
                raise RuntimeError(f"Load failed for {ext}: shape={loaded.shape}, columns={loaded.columns.tolist()}")
            if fast_row_count(path) != 2:
                raise RuntimeError(f"Row count failed for {ext}: {fast_row_count(path)}")

        ods_path = tmp / "sample.ods"
        data.to_excel(ods_path, engine="odf", index=False)
        if list_sheets(ods_path) != ["Sheet1"]:
            raise RuntimeError(f"Sheet listing failed for .ods: {list_sheets(ods_path)}")
        if load_excel(ods_path).shape[0] != 2:
            raise RuntimeError("Load failed for .ods")
        if fast_row_count(ods_path) != 2:
            raise RuntimeError(f"Row count failed for .ods: {fast_row_count(ods_path)}")

        tsv_path = tmp / "sample.tsv"
        data.to_csv(tsv_path, sep="\t", index=False)
        if list_sheets(tsv_path) != ["Delimited Data"]:
            raise RuntimeError(f"Sheet listing failed for .tsv: {list_sheets(tsv_path)}")
        if load_excel(tsv_path).shape[0] != 2:
            raise RuntimeError("Load failed for .tsv")
        if fast_row_count(tsv_path) != 2:
            raise RuntimeError(f"Row count failed for .tsv: {fast_row_count(tsv_path)}")

    print("OK Loader supports extended Excel formats")


def main() -> None:
    try:
        with TemporaryDirectory(prefix="excelmatcher-pro-test-") as tmpdir:
            app = create_app(Path(tmpdir))
            print("OK App created successfully")

            routes = {route.path for route in app.routes}
            required = {
                "/health",
                "/api/load-file",
                "/api/get-preview",
                "/api/load-full/{file_id}",
                "/api/ai/suggest-mappings",
                "/api/ai/tolerances",
                "/api/ai/prediction",
                "/api/run-match",
                "/api/match-status/{job_id}",
                "/api/match-result/{job_id}",
                "/api/user-correction",
                "/api/ai/stats",
            }

            for route in sorted(required):
                if route in routes:
                    print(f"OK {route}")
                else:
                    print(f"FAIL {route} NOT FOUND")
                    raise RuntimeError(f"Missing route: {route}")

            run_split_smoke_test()
            run_invoice_padding_smoke_test()
            run_qty_shortage_remark_smoke_test()
            run_key_mismatch_not_qty_smoke_test()
            run_blank_barcode_split_value_smoke_test()
            run_qty_mismatch_value_detail_smoke_test()
            run_summary_export_smoke_test()
            run_loader_format_support_smoke_test()
            print("\nOK Backend validation complete!")
    except Exception as exc:
        print(f"FAIL Backend validation failed: {exc}")
        raise


if __name__ == "__main__":
    main()
